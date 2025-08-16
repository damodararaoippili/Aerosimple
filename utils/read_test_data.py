import openpyxl
def read_login_data(sheet_name):
    file = r"C://Users//damod//PycharmProjects//AerosimpleWebApp//data//Aerosimple.xlsx"
    workbook = openpyxl.load_workbook(file)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Sheet name: {sheet_name}")
        test_data = {}
        for row in range(1, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key is not None:
                test_data[key] = value
        return test_data
    else:
        print(f"Sheet '{sheet_name}' not found in workbook.")

def read_test_case_data(sheet_name, testcase_ID):
    file = r"C://Users//damod//PycharmProjects//AerosimpleWebApp//data//Aerosimple.xlsx"
    workbook = openpyxl.load_workbook(file)

    if sheet_name not in workbook.sheetnames:
        print(f" Sheet '{sheet_name}' not found.")
        return None

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    for row in range(2, max_row + 1):
        testcase_cell_value = sheet.cell(row=row, column=1).value
        if testcase_cell_value == testcase_ID:
            header_row = row - 1
            headers = [sheet.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]
            values = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
            test_data = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
            print(test_data)
            return test_data

    print(f"Test case ID '{testcase_ID}' not found.")
    return None


